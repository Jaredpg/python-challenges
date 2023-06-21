import FlaskUI
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask

app = Flask(__name__)

@app.route("/")
def hello():  
    return render_template('index.html')

@app.route("/home", methods=['GET'])
def home(): 
    return render_template('some_page.html')


#wb = Workbook()
#ws = wb.active
#ws['A1'] = "birthday"
#ws['B1'] = "fav colour"
#ws['C1'] = "favourite food"
#ws.append(["1st Jan 2008", "green", "pretzels"])
#ws.append(["1", "2", "3"])

#wb.save("FlaskUI.xlsx")

#workbook = load_workbook(filename="FlaskUI.xlsx")
#workbook.sheetnames

#sheet = workbook.active
#sheet.title

#wb.save("FlaskUI.xlsx")



