import os
from openpyxl import Workbook
import sys


print("press anything to start: ")
try:
    records_start = input("")
    while True:
        if records_start == "":
            records_start = input("please enter something: ")
            print("")
        else:
            break
except NameError:
    print("Error: program ended")

def return_answer():
    print("")
    answer = input("press 1 to go back or 2 to exit: ")
    if answer == "1":
        print("you have chosen to go back ")
        print("")
        records_start = True
    elif answer == "2":
        print("program exited")
        sys.exit()
    elif answer == "":
        print("error - answer has to be 1 or 2: ")
        return_answer()
    else:
        print("error: answer has to be 1 or 2")
        return_answer()
        
def record_make():
    file_name = "project-records/" + input('Filename with extension, e.g. example.txt: ')
    with open(file_name, 'w', encoding='utf-8') as my_file:
        r_name = input(" name: ")
        r_age = input(" age: ")
        r_favourite_colour = input(" favourite colour: ")
        r_hobbies = input(" hobbies: ")
        r_pets = input(" pets: ")
        my_file.write(' name: ' + r_name)
        my_file.write(' age: ' + r_age)
        my_file.write(' favourite colour: ' + r_favourite_colour)
        my_file.write(' hobbies: ' + r_hobbies)
        my_file.write(' hobbies: ' + r_pets)
        print("record has been saved into library")

def delete_option():
    print("are you sure you would like to delete this file?")
    print("1 = yes, 2 = no")
    delete_option = input("")
    if delete_option == "1":
        print("deletion rejected: please don't delete pre-defined records")
        return_answer()        
    elif delete_option == "2":
        print("deletion has been cancelled")
        return_answer()

#class file_manager:
    #def__init__(self, create_file, delete_file, cancel_file):

    #create_file = 

    #delete_file = 

    #cancel_file =

#def pd_record():
    #print("accessing pre-defined record...")
    #print("")
    #record_option = int(input())
    #match record_option:
        #case 1:
            #pd_record_name = birthday
        #case 2:
            #pd_record_name = favfoods
        #case 3:
            #pd_record_name = pets
    #pdr_name = open("project-records/" + pd_record_name + ".txt", "r")
    #print(pdr_name.read())
    #return_answer()

if __name__ == "__main__":
    path = "/"
    while records_start:
        print("this is my record management project: ")
        print("")
        print("1 - access records")
        print("2 - create records")
        print("3 - delete records")
        print("4 - file manager")
        print("5 - exit program")

        records_start = input("")
        if records_start == "1":
            print("you have chosen to access records: ")
            print("")
            print("1 - pre-defined records")
            print("2 - saved records")
            print("3 - cancel action")
            access_record = (input(""))

            if access_record == "1":
                print("which records would you like to access...")
                print("")
                print("1 - birthday")
                print("2 - favourite food")
                print("3 - pets")
                print("4 - cancel action")

                record_option = input("")

                if record_option == "1":
                    print("accessing birthday records...")
                    print("")
                    birthday = open("project-records/birthday.txt", "r")
                    print(birthday.read())
                    return_answer()

                elif record_option == "2":
                    print("accessing favourite foods records...")
                    print("")
                    favfoods = open("project-records/favouritefoods.txt", "r")
                    print(favfoods.read())
                    return_answer()

                elif record_option == "3":
                    print("accessing pet records...")
                    print("")
                    pets = open("project-records/pets.txt", "r")
                    print(pets.read())
                    return_answer()

                elif record_option == "4":
                    print("action cancelled")
                    return_answer()
                
                else:
                    print("error: that is not a option")
                    return_answer()

            elif access_record == "2":
                print("accessing saved records...")
                print("")
                file_list = os.listdir("project-records")
                print("current files saved: ")
                print("")
                for file in file_list:
                    print(file)
                print("")
                access_saved_record = input("please input the record you would like to access: ")
                try:
                    saved_record = open("project-records/" + access_saved_record, "r")
                    print("")
                    print(saved_record.read())
                    return_answer()
                except NameError:
                    ("Error: file could not be found")
                    return_answer()
                
            elif access_record == "3":
                print("action cancelled")
                return_answer()

            else:
                print("error: that is not a option")
                return_answer()

        elif records_start == "2":
            print("you have chosen to create records! ")
            print("")
            print("1 - create a new file")
            print("2 - create basic record")
            print("3 - input user data into xlsx file")
            print("4 - cancel create")

            reply = input("")
            if reply == "1":
                print("write what you would like to save: ")

                file_name = "project-records/" + input('Filename with extension, e.g. example.txt: ')
                with open(file_name, 'w', encoding='utf-8') as my_file:
                    my_file.write(input('Your message: '))

                print("file has been saved into library")
                return_answer()
            
            elif reply == "2":
                record_make()
                return_answer()

            elif reply == "3":
                name = input("write down your full name: ")
                birthday = input("write down your birthday: ")
                favcolour = input("write down your favourite colour: ")
                hobby = input("write down a hobby: ")
                print("data has been save")
                print("")
                print("name: " + name)
                print("birthday: " + birthday)
                print("favourite colour: " + favcolour)
                print("hobby: " + hobby)

                wb = Workbook()
                ws = wb.active
                ws['A1'] = "name"
                ws['B1'] = "birthday"
                ws['C1'] = "favourite colour"
                ws['D1'] = "hobby"
                ws.append([name, birthday, favcolour, hobby])
                user_filename = input("insert file name: ")
                wb.save("project-records/" + user_filename + ".xlsx")
                return_answer()

            elif reply == "4":
                print("creation cancelled")
                return_answer()
            else:
                print("error: that is not a option")
                return_answer()

        elif records_start == "3":
            print("you have chosen to delete records: ")
            print("what record would you like to delete? ")
            print("")
            print("1 - pre-defined files")
            print("2 - enter file")
            print("3 - cancel deletion")

            delete_record = input("")

            if delete_record == "1": 
                print("")
                print("1 - birthday")
                print("2 - favourite foods")
                print("3 - pets")
                print("4 - cancel deletion")
                predefined_delete = input("")
                if predefined_delete == "1":
                    delete_option()
                elif predefined_delete == "2":
                    delete_option()    
                elif predefined_delete == "3":
                    delete_option()
                elif predefined_delete == "4":
                    print("deletion has been cancelled")
                    return_answer()
                else:
                    print("error: that is not a option")
                    return_answer()

            elif delete_record == "2":
                file_list = os.listdir("project-records")
                print("current files saved: ")
                print("")
                for file in file_list:
                    print(file)
                print("")

                input_delete_record = input("input the file you would like to delete: ")
                print("looking for file...")
                print("are you sure you would like to delete this file?")
                print("1 = yes, 2 = no: ")
                delete_option = input("")
                
            if delete_option == "1":
                try:
                    deletion = os.remove("project-records/" + input_delete_record)
                    print("file has been deleted")
                    return_answer()
                except NameError:
                    print("Error: file could not be found")
                    return_answer()
    
            elif delete_option == "2":
                print("deletion has been cancelled")
                return_answer()
            
            else:
                print("error: that is not a option")
                return_answer()

        elif records_start == "4":
            print("accessing file manager...")
            file_list = os.listdir("project-records")
            print("current files saved: ")
            print("")
            for file in file_list:
                print(file)
            print("")
            file_m = input("input file you would like to edit: ")
            print("you have selected: " + file_m)
            print("access, delete, cancel action")
            file_decision = input("what would you like to do with this file?: ")

            if file_decision == "access":
                try:
                    file_m = open("project-records/" + file_m, "r")
                    print("")
                    print(file_m.read())
                    return_answer()
                except NameError:
                    ("Error: file could not be found")
                    return_answer()
            elif file_decision == "delete":
                print("are you sure you would like to delete this file?")
                print("1 = yes, 2 = no: ")
                delete_option = input("")
                
                if delete_option == "1":
                    try:
                        os.remove("project-records/" + file_m)
                        print("file has been deleted")
                        return_answer()
                    except NameError:
                        print("Error: file could not be found")
                        return_answer()
        
                elif delete_option == "2":
                    print("deletion has been cancelled")
                    return_answer()

                else:
                    print("error: that is not a option")
                    return_answer()

            elif file_decision =="cancel":
                print("action has been cancelled")
                return_answer

            elif access_record == "3":
                print("action cancelled")
                return_answer()
            else:
                print("Error - that wasn't an option")
                return_answer()

        elif records_start == "5":
            print("program has been exited")
            break
        else:
            print("error: that is not a option")
            return_answer()
            records_start = True
