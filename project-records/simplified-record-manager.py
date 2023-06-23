import os
from openpyxl import Workbook
import sys

print("Enter something to start: ")
try:
    records_start = input("")
    while True:
        if records_start == "":
            records_start = input("please enter something: ")
            print("")
        else:
            break
except NameError:
    print("Error: program ended")

def return_answer():
    print("")
    answer = input("press 1 to go back or 2 to exit: ")
    if answer == "1":
        print("you have chosen to go back ")
        print("")
        records_start = True
    elif answer == "2":
        print("program exited")
        sys.exit()
    elif answer == "":
        print("error - answer has to be 1 or 2: ")
        return_answer()
    else:
        print("error: answer has to be 1 or 2")
        return_answer()
        
def record_make():
    file_name = "project-records/" + input('Filename with extension, e.g. example.txt: ')
    with open(file_name, 'w', encoding='utf-8') as my_file:
        r_name = input(" name: ")
        r_age = input(" age: ")
        r_favourite_colour = input(" favourite colour: ")
        r_hobbies = input(" hobbies: ")
        r_pets = input(" pets: ")
        my_file.write(' name: ' + r_name)
        my_file.write(' age: ' + r_age)
        my_file.write(' favourite colour: ' + r_favourite_colour)
        my_file.write(' hobbies: ' + r_hobbies)
        my_file.write(' hobbies: ' + r_pets)
        print("record has been saved into library")

def delete_option():
    print("are you sure you would like to delete this file?")
    print("1 = yes, 2 = no")
    delete_option = input("")
    if delete_option == "1":
        print("deletion rejected: please don't delete pre-defined records")
        return_answer()        
    elif delete_option == "2":
        print("deletion has been cancelled")
        return_answer()

if __name__ == "__main__":
    path = "/"
    while records_start:
        print("this is my record management project: ")
        print("")
        print(" 1 - create records")
        print(" 2 - file manager")
        print(" 3 - exit program")
        print("")
        records_start = input("")

        if records_start == "1":
            print("you have chosen to create records! ")
            print("")
            print(" 1 - create a new file")
            print(" 2 - create basic record")
            print(" 3 - input user data into xlsx file")
            print(" 4 - cancel create")
            print("")

            reply = input("")
            if reply == "1":
                print("write what you would like to save: ")

                file_name = "project-records/" + input('Filename with extension, e.g. example.txt: ')
                with open(file_name, 'w', encoding='utf-8') as my_file:
                    my_file.write(input('Your message: '))

                print("file has been saved into library")
                return_answer()
            
            elif reply == "2":
                record_make()
                return_answer()

            elif reply == "3":
                name = input("write down your full name: ")
                birthday = input("write down your birthday: ")
                favcolour = input("write down your favourite colour: ")
                hobby = input("write down a hobby: ")
                print("data has been save")
                print("")
                print(" name: " + name)
                print(" birthday: " + birthday)
                print(" favourite colour: " + favcolour)
                print(" hobby: " + hobby)

                wb = Workbook()
                ws = wb.active
                ws['A1'] = "name"
                ws['B1'] = "birthday"
                ws['C1'] = "favourite colour"
                ws['D1'] = "hobby"
                ws.append([name, birthday, favcolour, hobby])
                user_filename = input("insert file name: ")
                wb.save("project-records/" + user_filename + ".xlsx")
                return_answer()

            elif reply == "4":
                print("creation cancelled")
                return_answer()
            else:
                print("error: that is not a option")
                return_answer()

        elif records_start == "2":
            print("accessing file manager...")
            file_list = os.listdir("project-records")
            print("current files saved: ")
            print("")
            for file in file_list:
                print(file)
            print("")
            file_m = input("input file you would like to edit: ")
            if file_m == "":
                print("Error: you have not selected a file")
                return_answer()
            elif file_m == False:
                print("Error: you have not selected a file")
                return_answer()
            print("")
            print("you have selected: " + file_m)
            print("access, delete, cancel")
            file_decision = input("what would you like to do with this file?: ")

            if file_decision == "access":
                try:
                    file_m = open("project-records/" + file_m, "r")
                    print("")
                    print(file_m.read())
                    return_answer()
                except NameError:
                    ("Error: file could not be found")
                    return_answer()
                    records_start = True
            elif file_decision == "delete":
                print("are you sure you would like to delete this file?")
                print("1 = yes, 2 = no: ")
                delete_option = input("")
                
                if delete_option == "1":
                    try:
                        os.remove("project-records/" + file_m)
                        print("file has been deleted")
                        return_answer()
                    except NameError:
                        print("Error: file could not be found")
                        return_answer()
        
                elif delete_option == "2":
                    print("deletion has been cancelled")
                    return_answer()

                else:
                    print("error: that is not a option")
                    return_answer()

            elif file_decision =="cancel":
                print("action has been cancelled")
                return_answer()

            else:
                print("Error - that wasn't an option")
                return_answer()

        elif records_start == "3":
            print("program has been exited")
            break
        else:
            print("error: that is not a option")
            return_answer()
            records_start = True
